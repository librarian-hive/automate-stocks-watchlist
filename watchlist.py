import yfinance as yf
import datetime as dt
import openpyxl as xl
import os
from pandas_datareader import data as pdr
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, Alignment, Border, Side, colors 
import pprint

def get_xl_file_path(xl_file):
    path = os.getcwd() + '/Desktop/'
    file_path = path + xl_file
    return file_path

def get_dates(today):
    if today == 'auto':
        today = dt.datetime.today()
    dates = []
    date = today.date()
    weekday = date.isoweekday()
    if weekday > 6:
        friday = date - dt.timedelta(2)
        dates.append(friday)
        return dates
    sunday = date - dt.timedelta(weekday)
    friday = sunday - dt.timedelta(2)
    holidays=[dt.date(2020,7,3), dt.date(2020,12,25), dt.date(2021,1,1), dt.date(2021,4,2), 
              dt.date(2021,12,24), dt.date(2022,4,15)]
    while friday in holidays:
        friday = friday - dt.timedelta(1)
    dates.append(friday)
    count = 1
    while weekday > 0 and count < 6:
        dates.append(sunday + dt.timedelta(count))
        count += 1
        weekday -=1
    return dates

def build_xl_file_worksheet(file_path, stocks, dates, strategy, position):
    wb = xl.load_workbook(file_path)
    sheet_name = str(dates[0])
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    title_row = 'Trading Week Of ' + str(dates[0]) + ' - ' + str(dates[len(dates)-1])
    header_row = ['Stocks','Strategy/TF', 'Position']
    for date in dates:
        header_row.append(str(date))
    row = 2
    col = 2
    col_letter1 = get_column_letter(col)
    col_letter2 = get_column_letter(col + 8)
    cell1 = col_letter1 + str(row)
    cell2 = col_letter2 + str(row)
    cells_merged = cell1 + ':' + cell2
    sheet.merge_cells(cells_merged)
    sheet[cell1] = title_row
    row = 3
    col = 2
    for header_cell in header_row:
        col_letter = get_column_letter(col)
        cell = col_letter + str(row)
        sheet[cell] = header_cell
        col += 1
    row = 4
    col = 2
    yf.pdr_override()
    index = 0
    for stock in stocks:
        col_letter = get_column_letter(col)
        cell = col_letter + str(row)
        sheet[cell] = stock
        col += 1
        col_letter = get_column_letter(col)
        cell = col_letter + str(row)
        sheet[cell] = strategy[index]
        col += 1
        col_letter = get_column_letter(col)
        cell = col_letter + str(row)
        sheet[cell] = position[index]
        col += 1
        df = pdr.get_data_yahoo(stock, dates[0], dates[len(dates)-1]+dt.timedelta(1))
        for close in df['Adj Close']:
            col_letter = get_column_letter(col)
            cell = col_letter + str(row)
            price = round(close, 2)
            sheet[cell] = price
            col += 1
        row += 1
        col = 2
        index += 1
    wb.save(file_path)
    return sheet_name

def apply_syles_xl(file_path, sheet_name, last_row, date_columns):
    wb = xl.load_workbook(file_path)
    sheet = wb[sheet_name]
    font_24 = Font(size=24)
    font_20 = Font(size=20)
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    gray_fill = PatternFill(start_color='CCCCCCCC', end_color='CCCCCCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right')
    sheet['B2'].font = font_24
    sheet['B2'].fill = gray_fill
    sheet['B2'].alignment = align_center

    for row in sheet['B3:J3']:
        for cell in row:
            cell.font = font_20
            cell.fill = gray_fill
            cell.alignment = align_center
        
    for row in sheet['B4:D' + last_row]:
        for cell in row:
            cell.font = font_20
            cell.fill = gray_fill
            cell.alignment = align_center
        
    for row in sheet['E4:E' + last_row]:
        for cell in row:
            cell.font = font_20
            cell.fill = yellow_fill
            cell.alignment = align_right

    for num in range(date_columns):
        col = 6 + num 
        col_letter = get_column_letter(col)
        for i in range(int(last_row)-3):
            try:
                cell_name = col_letter + str(i+4)
                cell = sheet[cell_name]
                friday_price_cell = sheet['E' + str(i+4)]
                cell.font = font_20
                if cell.value > friday_price_cell.value:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            except:
                continue

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell == sheet['B2']:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.5
        sheet.column_dimensions[column].width = adjusted_width

    wb.save(file_path)

def auto_list(stocks, text):
    text_list=[]
    for i in range(len(stocks)):
        text_list.append(text)
    return text_list

if __name__ == '__main__':
    stocks = ['CGNX','ZS','NET','EBAY','ETSY','STAA','TDOC','DS','WMT','PENN']
    #strategy = ['StockDweebs','StockDweebs','StockDweebs','StockDweebs','StockDweebs',
    #            'StockDweebs','StockDweebs','StockDweebs','StockDweebs','StockDweebs']
    strategy = auto_list(stocks, 'StockDweeb')
    #position = ['LONG','LONG','LONG','LONG','LONG','LONG','LONG','LONG','LONG','LONG']
    position = auto_list(stocks, 'LONG')
    #today = dt.datetime(2020,6,20)
    today = 'auto'
    dates = get_dates(today)
    file_path = get_xl_file_path('watchlist.xlsx')
    sheet_name = build_xl_file_worksheet(file_path, stocks, dates, strategy, position)
    last_row = str(len(stocks) + 3)
    date_columns = len(dates) - 1
    apply_syles_xl(file_path, sheet_name, last_row, date_columns)
